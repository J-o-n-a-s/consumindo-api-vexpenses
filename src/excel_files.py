from datetime import date
from os.path import isfile
from time import sleep
from tkinter import filedialog

from file_header import LINE_SIZE
from functions import format_print
from openpyxl import Workbook, load_workbook


def save_file(path: str, rows: list) -> None:
    file_is_open = False
    workbook = None
    file_name = path + '/' + date.today().strftime('%Y_%m_%d') + '.xlsx'

    if isfile(file_name):
        try:
            workbook = load_workbook(file_name)
        except Exception as error:
            format_print(
                fill_char=' ',
                line_size=LINE_SIZE,
                text=f'Ocorreu erro ao abrir o arquivo {error.__class__}.',
            )

            if 'error' in locals():
                del error

        else:
            file_is_open = True

    else:
        try:
            workbook = Workbook()
        except Exception as error:
            format_print(
                fill_char=' ',
                line_size=LINE_SIZE,
                text=f'Ocorreu erro ao criar o arquivo {error.__class__}.',
            )

            if 'error' in locals():
                del error

        else:
            file_is_open = True

    if file_is_open:
        del file_is_open
        worksheet = workbook.active

        for row in rows:
            try:
                worksheet.append(row[:])
                workbook.save(file_name)
            except Exception as error:
                format_print(
                    fill_char=' ',
                    line_size=LINE_SIZE,
                    text=f'Ocorreu erro ao salvar o arquivo {error.__class__}.',
                )

                if 'error' in locals():
                    del error

            else:
                format_print(
                    fill_char=' ',
                    line_size=LINE_SIZE,
                    text=f'Arquivo "{file_name}" salvo com sucesso.',
                )

        if 'row' in locals():
            del row

        if 'worksheet' in locals():
            del worksheet

        format_print(fill_char='-', line_size=LINE_SIZE, text='')

    if 'file_name' in locals():
        del file_name

    if 'workbook' in locals():
        del workbook


def select_path() -> str:
    format_print(fill_char='-', line_size=LINE_SIZE, text='')
    format_print(
        fill_char=' ',
        text='Por gentileza, selecione o diretório para a criação dos arquivos.',
        line_size=LINE_SIZE,
    )
    sleep(1)
    path = filedialog.askdirectory()
    format_print(
        fill_char=' ',
        line_size=LINE_SIZE,
        text=f'Diretório selecionado "{path.replace("/", "\\")}"',
    )
    format_print(fill_char='-', line_size=LINE_SIZE, text='')

    return path
